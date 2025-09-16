# import time
# import json
# import argparse
# from pathlib import Path
# from datetime import datetime
# from selenium import webdriver
# from selenium.webdriver.chrome.service import Service
# from selenium.webdriver.chrome.options import Options
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# from webdriver_manager.chrome import ChromeDriverManager
# from openpyxl import Workbook, load_workbook

# APOLLO_URL = "https://app.apollo.io/"
# ACCOUNT_JSON_PATH = Path.cwd() / "accounts.json"
# EXCEL_FILE = Path(r"C:/Users/X1/Selenium/apollo/new_a.xlsx")

# def get_profile_path(account_name: str) -> str:
#     if not ACCOUNT_JSON_PATH.exists():
#         raise FileNotFoundError(f"{ACCOUNT_JSON_PATH} not found")
#     with open(ACCOUNT_JSON_PATH, "r") as f:
#         accounts = json.load(f)
#     if account_name not in accounts:
#         raise ValueError(f"Account '{account_name}' not found in {ACCOUNT_JSON_PATH}")
#     return accounts[account_name]

# def create_driver_with_profile(profile_path: str):
#     options = Options()
#     options.add_argument("--start-maximized")
#     options.add_argument("--disable-blink-features=AutomationControlled")
#     options.add_argument(f"--user-data-dir={profile_path}")
#     options.add_experimental_option("detach", True)
#     options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
#     options.add_experimental_option("useAutomationExtension", False)
#     service = Service(ChromeDriverManager().install())
#     return webdriver.Chrome(service=service, options=options)

# def init_excel():
#     headers = ["First Name", "Last Name", "LinkedIn URL",
#                "View Profile ✔️", "First Task Done ✔️",
#                "Browser User", "Timestamp"]
#     if EXCEL_FILE.exists():
#         wb = load_workbook(EXCEL_FILE)
#         ws = wb.active
#         if ws.max_row < 1 or ws[1][0].value != headers[0]:
#             ws.insert_rows(1)
#             for col, header in enumerate(headers, start=1):
#                 ws.cell(row=1, column=col).value = header
#     else:
#         wb = Workbook()
#         ws = wb.active
#         ws.append(headers)
#     return wb, ws

# def log_to_excel(ws, first_name, last_name, linkedin_url,
#                  view_profile_done=False, first_task_done=False,
#                  browser_user=""):
#     timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
#     ws.append([
#         first_name,
#         last_name,
#         linkedin_url,
#         "✔️" if view_profile_done else "",
#         "✔️" if first_task_done else "",
#         browser_user,
#         timestamp
#     ])

# def save_excel_with_retry(wb, max_retries=3):
#     for attempt in range(max_retries):
#         try:
#             wb.save(EXCEL_FILE)
#             print(f"Excel saved successfully on attempt {attempt + 1}")
#             return
#         except PermissionError as e:
#             print(f"Permission denied on attempt {attempt + 1}: {e}. Retrying in 2 seconds...")
#             time.sleep(2)
#     raise PermissionError(f"Failed to save Excel after {max_retries} attempts. Ensure the file is not open elsewhere.")

# def pause_for_linkedin_verification(driver):
#     print("\nLinkedIn verification may be required. Please complete verification manually in the browser.")
#     input("Once done, press Enter to continue...")

# def main():
#     parser = argparse.ArgumentParser()
#     parser.add_argument("--account", type=str, required=True, help="Account name in accounts.json")
#     args = parser.parse_args()
#     account_name = args.account

#     try:
#         profile_path = get_profile_path(account_name)
#     except Exception as e:
#         print("Error:", e)
#         return

#     driver = create_driver_with_profile(profile_path)
#     wait = WebDriverWait(driver, 30)
#     wb, ws = init_excel()

#     try:
#         driver.get(APOLLO_URL)
#         time.sleep(5)

#         # Click Sequences
#         sequences_link = wait.until(EC.element_to_be_clickable((By.ID, "side-nav-sequences")))
#         sequences_link.click()
#         print("Clicked 'Sequences'")
#         time.sleep(5)

#         # Click NGS Campaign
#         ngs_campaign = wait.until(EC.element_to_be_clickable(
#             (By.XPATH, "//span[contains(text(),'NGS Campaign')]")
#         ))
#         ngs_campaign.click()
#         print("Clicked 'NGS Campaign'")
#         time.sleep(5)

#         # Click Tasks tab
#         tasks_tab = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[contains(text(),'Tasks')]")))
#         tasks_tab.click()
#         print("Clicked 'Tasks'")
#         time.sleep(3)

#         rows = driver.find_elements(By.CSS_SELECTOR, "div[id^='table-row-']")
#         max_records = min(5, len(rows))

#         for index in range(max_records):
#             # Re-fetch rows after each iteration to handle any page changes
#             rows = driver.find_elements(By.CSS_SELECTOR, "div[id^='table-row-']")
#             if index >= len(rows):
#                 print(f"Row {index} no longer exists, skipping...")
#                 continue
                
#             row = rows[index]

#             # Initialize variables
#             linkedin_url = "N/A"
#             first_name = "Unknown"
#             last_name = "Unknown"
#             first_task_done = False
            
#             # Store initial number of tabs
#             initial_tabs = len(driver.window_handles)

#             # --------------------------
#             # Execute Apollo task (this should open LinkedIn in new tab)
#             # --------------------------
#             try:
#                 execute_btn = row.find_element(By.CSS_SELECTOR, "button[aria-label='Execute task']")
#                 driver.execute_script(
#                     "arguments[0].scrollIntoView({behavior: 'smooth', block: 'center', inline: 'center'});",
#                     execute_btn)
#                 time.sleep(1)
#                 execute_btn.click()
#                 first_task_done = True
#                 print("Execute task button clicked!")

#                 # Wait for new tab to open (LinkedIn should open automatically)
#                 time.sleep(5)
                
#                 # Check if a new tab opened
#                 current_tabs = len(driver.window_handles)
#                 if current_tabs > initial_tabs:
#                     print("New tab detected - switching to LinkedIn tab...")
                    
#                     # Switch to the new tab (LinkedIn)
#                     driver.switch_to.window(driver.window_handles[-1])
#                     time.sleep(5)  # Wait for LinkedIn to load
                    
#                     # Handle LinkedIn verification if needed
#                     if "checkpoint" in driver.current_url or "verify" in driver.current_url:
#                         pause_for_linkedin_verification(driver)
                    
#                     try:
#                         # Wait for LinkedIn to load properly
#                         WebDriverWait(driver, 20).until(EC.title_contains("LinkedIn"))
                        
#                         # Get the actual LinkedIn URL
#                         linkedin_url = driver.current_url
#                         print(f"LinkedIn URL captured: {linkedin_url}")
                        
#                         # Get profile name from LinkedIn
#                         try:
#                             full_name_elem = WebDriverWait(driver, 15).until(
#                                 EC.presence_of_element_located((By.XPATH, "//h1[contains(@class, 'text-heading-xlarge')]"))
#                             )
#                             full_name = full_name_elem.text.strip()
#                             name_parts = full_name.split(" ", 1)
#                             first_name = name_parts[0] if name_parts else "Unknown"
#                             last_name = name_parts[1] if len(name_parts) > 1 else "Unknown"
#                             print(f"Name extracted: {first_name} {last_name}")
#                         except Exception as e:
#                             print(f"Could not extract name: {e}")
                        
#                     except Exception as e:
#                         print(f"LinkedIn page processing failed: {e}")
#                         # Still use the URL even if name extraction fails
#                         linkedin_url = driver.current_url
                    
#                     # Close LinkedIn tab and switch back to Apollo
#                     driver.close()
#                     driver.switch_to.window(driver.window_handles[0])
                    
#                 else:
#                     print("No new tab opened after clicking execute task")

#                 # Get Apollo URL after task execution
#                 time.sleep(2)
#                 apollo_url = driver.current_url
#                 print("Apollo URL after task:", apollo_url)

#             except Exception as e:
#                 print(f"Execute task failed: {e}")

#             # Log to Excel if we got LinkedIn info
#             if linkedin_url != "N/A":
#                 log_to_excel(ws, first_name, last_name, linkedin_url,
#                              view_profile_done=True, first_task_done=first_task_done,
#                              browser_user=account_name)
#                 save_excel_with_retry(wb)
#                 print(f"Logged to Excel: {first_name} {last_name} - {linkedin_url}")
#             else:
#                 print("No LinkedIn URL captured for this task")

#             # Wait 2 minutes before next row
#             print("Waiting 2 minutes before processing the next person...")
#             time.sleep(120)

#         print("Completed tasks for all records.")

#     except Exception as e:
#         print("Error during automation:", e)

#     finally:
#         driver.quit()
#         print("Browser closed.")

# if __name__ == "__main__":
#     main()



import time
import json
import argparse
from pathlib import Path
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook


APOLLO_URL = "https://app.apollo.io/"
ACCOUNT_JSON_PATH = Path.cwd() / "accounts.json"
EXCEL_FILE = Path(r"C:/Users/X1/Selenium/apollo/new_a.xlsx")


def get_profile_path(account_name: str) -> str:
    if not ACCOUNT_JSON_PATH.exists():
        raise FileNotFoundError(f"{ACCOUNT_JSON_PATH} not found")
    with open(ACCOUNT_JSON_PATH, "r") as f:
        accounts = json.load(f)
    if account_name not in accounts:
        raise ValueError(f"Account '{account_name}' not found in {ACCOUNT_JSON_PATH}")
    return accounts[account_name]


def create_driver_with_profile(profile_path: str):
    options = Options()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument(f"--user-data-dir={profile_path}")
    options.add_experimental_option("detach", True)
    options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
    options.add_experimental_option("useAutomationExtension", False)
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)


def init_excel():
    headers = ["First Name", "Last Name", "LinkedIn URL",
               "View Profile ✔️", "First Task Done ✔️",
               "Browser User", "Timestamp"]
    if EXCEL_FILE.exists():
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        if ws.max_row < 1 or ws[1][0].value != headers[0]:
            ws.insert_rows(1)
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col).value = header
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
    return wb, ws


def log_to_excel(ws, first_name, last_name, linkedin_url,
                 view_profile_done=False, first_task_done=False,
                 browser_user=""):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([
        first_name,
        last_name,
        linkedin_url,
        "✔️" if view_profile_done else "",
        "✔️" if first_task_done else "",
        browser_user,
        timestamp
    ])


def save_excel_with_retry(wb, max_retries=3):
    for attempt in range(max_retries):
        try:
            wb.save(EXCEL_FILE)
            print(f"Excel saved successfully on attempt {attempt + 1}")
            return
        except PermissionError as e:
            print(f"Permission denied on attempt {attempt + 1}: {e}. Retrying in 2 seconds...")
            time.sleep(2)
    raise PermissionError(f"Failed to save Excel after {max_retries} attempts. Ensure the file is not open elsewhere.")


def pause_for_linkedin_verification(driver):
    print("\nLinkedIn verification may be required. Please complete verification manually in the browser.")
    input("Once done, press Enter to continue...")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--account", type=str, required=True, help="Account name in accounts.json")
    args = parser.parse_args()
    account_name = args.account


    try:
        profile_path = get_profile_path(account_name)
    except Exception as e:
        print("Error:", e)
        return


    driver = create_driver_with_profile(profile_path)
    wait = WebDriverWait(driver, 30)
    wb, ws = init_excel()


    try:
        driver.get(APOLLO_URL)
        time.sleep(5)


        # Click Sequences
        sequences_link = wait.until(EC.element_to_be_clickable((By.ID, "side-nav-sequences")))
        sequences_link.click()
        print("Clicked 'Sequences'")
        time.sleep(5)


        # Click NGS Campaign
        ngs_campaign = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//span[contains(text(),'NGS Campaign')]")
        ))
        ngs_campaign.click()
        print("Clicked 'NGS Campaign'")
        time.sleep(5)


        # Click Tasks tab
        tasks_tab = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[contains(text(),'Tasks')]")))
        tasks_tab.click()
        print("Clicked 'Tasks'")
        time.sleep(3)


        rows = driver.find_elements(By.CSS_SELECTOR, "div[id^='table-row-']")
        max_records = min(5, len(rows))


        for index in range(max_records):
            # Re-fetch rows after each iteration to handle any page changes
            rows = driver.find_elements(By.CSS_SELECTOR, "div[id^='table-row-']")
            if index >= len(rows):
                print(f"Row {index} no longer exists, skipping...")
                continue
                
            row = rows[index]


            # Initialize variables
            linkedin_url = "N/A"
            first_name = "Unknown"
            last_name = "Unknown"
            first_task_done = False
            
            # Store initial number of tabs
            initial_tabs = len(driver.window_handles)


            # --------------------------
            # Execute Apollo task (this should open LinkedIn in new tab)
            # --------------------------
            try:
                execute_btn = row.find_element(By.CSS_SELECTOR, "button[aria-label='Execute task']")
                driver.execute_script(
                    "arguments[0].scrollIntoView({behavior: 'smooth', block: 'center', inline: 'center'});",
                    execute_btn)
                time.sleep(1)
                execute_btn.click()
                first_task_done = True
                print("Execute task button clicked!")


                # Wait for new tab to open (LinkedIn should open automatically)
                time.sleep(5)
                
                # Check if a new tab opened
                current_tabs = len(driver.window_handles)
                if current_tabs > initial_tabs:
                    print("New tab detected - switching to LinkedIn tab...")
                    
                    # Switch to the new tab (LinkedIn)
                    driver.switch_to.window(driver.window_handles[-1])
                    time.sleep(5)  # Wait for LinkedIn to load
                    
                    # Handle LinkedIn verification if needed
                    if "checkpoint" in driver.current_url or "verify" in driver.current_url:
                        pause_for_linkedin_verification(driver)
                    
                    try:
                        # Wait for LinkedIn to load properly
                        WebDriverWait(driver, 20).until(EC.title_contains("LinkedIn"))
                        
                        # Get the actual LinkedIn URL
                        linkedin_url = driver.current_url
                        print(f"LinkedIn URL captured: {linkedin_url}")
                        
                        # Get profile name from the Apollo overlay on LinkedIn
                        try:
                            # Wait for Apollo overlay to appear on LinkedIn
                            time.sleep(3)
                            
                            # Try to find the Apollo info div with person's name
                            name_elem = WebDriverWait(driver, 10).until(
                                EC.presence_of_element_located((By.XPATH, "//span[@class='x_ieGBp']"))
                            )
                            full_name = name_elem.text.strip()
                            name_parts = full_name.split(" ", 1)
                            first_name = name_parts[0] if name_parts else "Unknown"
                            last_name = name_parts[1] if len(name_parts) > 1 else "Unknown"
                            print(f"Name extracted from Apollo overlay: {first_name} {last_name}")
                        except Exception as e:
                            print(f"Could not extract name from Apollo overlay: {e}")
                            # Fallback to LinkedIn profile name
                            try:
                                full_name_elem = WebDriverWait(driver, 10).until(
                                    EC.presence_of_element_located((By.XPATH, "//h1[contains(@class, 'text-heading-xlarge')]"))
                                )
                                full_name = full_name_elem.text.strip()
                                name_parts = full_name.split(" ", 1)
                                first_name = name_parts[0] if name_parts else "Unknown"
                                last_name = name_parts[1] if len(name_parts) > 1 else "Unknown"
                                print(f"Name extracted from LinkedIn profile: {first_name} {last_name}")
                            except Exception as e2:
                                print(f"Could not extract name from LinkedIn either: {e2}")
                        
                    except Exception as e:
                        print(f"LinkedIn page processing failed: {e}")
                        # Still use the URL even if name extraction fails
                        linkedin_url = driver.current_url
                    
                    # Close LinkedIn tab and switch back to Apollo
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    
                else:
                    print("No new tab opened after clicking execute task")


                # Get Apollo URL after task execution
                time.sleep(2)
                apollo_url = driver.current_url
                print("Apollo URL after task:", apollo_url)


            except Exception as e:
                print(f"Execute task failed: {e}")


            # Log to Excel if we got LinkedIn info
            if linkedin_url != "N/A":
                log_to_excel(ws, first_name, last_name, linkedin_url,
                             view_profile_done=True, first_task_done=first_task_done,
                             browser_user=account_name)
                save_excel_with_retry(wb)
                print(f"Logged to Excel: {first_name} {last_name} - {linkedin_url}")
            else:
                print("No LinkedIn URL captured for this task")


            # Wait 2 minutes before next row
            print("Waiting 2 minutes before processing the next person...")
            time.sleep(120)


        print("Completed tasks for all records.")


    except Exception as e:
        print("Error during automation:", e)


    finally:
        driver.quit()
        print("Browser closed.")


if __name__ == "__main__":
    main()
